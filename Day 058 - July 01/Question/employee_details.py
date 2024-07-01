import datetime
import re
from openpyxl import Workbook, load_workbook
from fpdf import FPDF

def calculate_age(dob):
    today = datetime.date.today()
    return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))

def validate_email(email):
    regex = r'^\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b'
    return re.match(regex, email)

def validate_phone(phone):
    regex = r"^\d{10}$"
    return re.match(regex, phone)

def add_employee():
    while True:
        name = input("Enter Name: ")
        dob_str = input("Enter Date of Birth (YYYY-MM-DD): ")
        phone = input("Enter Phone Number: ")
        email = input("Enter Email: ")

        try:
            dob = datetime.datetime.strptime(dob_str, "%Y-%m-%d").date()
        except ValueError:
            print("Invalid Date of Birth format. Please enter in YYYY-MM-DD format.")
            continue

        if not validate_email(email):
            print("Invalid Email format.")
            continue

        if not validate_phone(phone):
            print("Invalid Phone Number format.")
            continue

        age = calculate_age(dob)
        return {"Name": name, "DOB": dob, "Phone": phone, "Email": email, "Age": age}


def save_as_excel():
    wb = load_workbook(filename="Day 058 - July 01/Question/employees.xlsx")
    ws = wb["Sheet1"]
    
    for employee in employees:
        ws.append([employee["Name"], employee["Age"], employee["DOB"], employee["Email"], employee["Phone"]])
    
    wb.save(filename="Day 058 - July 01/Question/employees.xlsx")
    wb.close()
    print("File saved successfully")

def save_as_pdf():
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)

    for emp in employees:
        pdf.cell(200, 10, txt=f"Name: {emp["Name"]}", ln=True)
        pdf.cell(200, 10, txt=f"DOB: {emp["Age"]}", ln=True)
        pdf.cell(200, 10, txt=f"Phone: {emp["DOB"]}", ln=True)
        pdf.cell(200, 10, txt=f"Email: {emp["Email"]}", ln=True)
        pdf.cell(200, 10, txt=f"Age: {emp["Phone"]}", ln=True)
        pdf.cell(200, 10, txt="", ln=True)

    pdf.output("Day 058 - July 01/Question/employees.pdf")
    print("File saved successfully")

def save_to_file():
    print(f"{'Save as':*^30}")
    print("1. Text file")
    print("2. Excel file")
    print("3. PDF file")
    print("Note: The filename will be \'employees\'")

    choice = int(input('Enter your choice: '))
    match choice:
        case 1:
            # text file
            with open('Day 058 - July 01/Question/employees.txt', 'a') as file:
                for employee in employees:
                    file.write(f"****Employee****\nName: {employee['Name']}\nAge: {employee['Age']}\nDate Of Birth: {employee['DOB']}\nEmail: {employee['Email']}\nPhone: {employee['Phone']}\n\n")
            
            print("File saved successfully")
            
        case 2:
            # excel file
            save_as_excel()
        case 3:
            # pdf file
            save_as_pdf()
        case _:
            print('Invalid choice. Please try again.')

def read_excel_data():
    wb = load_workbook(filename="Day 058 - July 01/Question/employees.xlsx", read_only=True)
    sheet = wb['Sheet1']

    sheet_data = []

    for index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if index == 1:
            continue
        
        row_data = {
            'Name': row[0],
            'Age': row[1],
            'DOB': row[2],
            'Email': row[3],
            'Phone': row[4]
        }
        sheet_data.append(row_data)

    wb.close()

    return sheet_data


# MAIN
employees = []
while True:
    print(f'{'Menu':*^30}')
    print('1. Add employee')
    print('2. Display all employees')
    print('3. Save to file')
    print('4. Read from file')
    print('5. Exit')
    
    choice = int(input('Enter your choice: '))

    match choice:
        case 1:
            # add employee
            new_employee = add_employee()
            employees.append(new_employee)
            print('Employee added successfully')
            print(f'{'':*^50}\n\n')
        case 2:
            # display employees
            count = 1
            for employee in employees:
                print(f"{f'Employee {count}':*^30}")
                count += 1
                print(f"Name: {employee["Name"]}\nAge: {employee["Age"]}\nDate Of Birth: {employee["DOB"]}\nEmail: {employee["Email"]}\nPhone: {employee["Phone"]}")
                print("\n")
            print(f'{'':*^50}\n\n') 
        case 3:
            # save to file
            save_to_file()
            print(f'{'':*^50}\n\n')
        case 4:
            # read from file
            xls_data = read_excel_data()
            print('File read successfully')
            print(f'{'':*^50}\n\n')
        case 5:
            break
        case _:
            print('Invalid choice. Please try again.')